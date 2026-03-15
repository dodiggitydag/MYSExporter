"""
merger.py - Key-based merge for MYS session proposals.

Merges fresh API data into an existing Excel workbook while preserving
ALL columns not present in the API response. The script only "owns"
columns whose header names match a key in api_records, plus two meta
columns (_SyncStatus, _LastSynced). Everything else - regardless of
where it sits in the sheet - is treated as human-created and never
modified.

Usage:
    from merger import merge_sessions
    merge_sessions(api_records: list[dict], output_path: str)
"""

import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

KEY_FIELD = "proposalguid"

# Starter human columns seeded ONLY on first run.
# Reviewers can rename, reorder, delete, or add more - the script won't
# touch any column whose header doesn't match an API field name.
SEED_HUMAN_COLUMNS = [
    {"name": "Rating", "width": 10},
    {"name": "Accept?", "width": 12},
    {"name": "Track/Category", "width": 20},
    {"name": "Reviewer Notes", "width": 40},
]

# Meta columns owned by the script (always written)
STATUS_COL = "_SyncStatus"
SYNCED_COL = "_LastSynced"
META_COLS = {STATUS_COL, SYNCED_COL}

# Styling
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
API_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HUMAN_HEADER_FILL = PatternFill("solid", fgColor="548235")
META_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
META_HEADER_FONT = Font(name="Arial", bold=True, color="666666", size=10)
REMOVED_FILL = PatternFill("solid", fgColor="FFC7CE")
NEW_FILL = PatternFill("solid", fgColor="C6EFCE")
NO_FILL = PatternFill()
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="D0D0D0"),
)

# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def merge_sessions(
    api_records: list[dict],
    output_path: str = "sessions.xlsx",
    key_field: str = KEY_FIELD,
    backup: bool = True,
) -> dict:
    """
    Merge fresh API data into an existing (or new) Excel workbook.

    Only columns whose header matches a key in api_records (or the two
    _Sync* meta columns) are written. All other columns are preserved
    exactly as-is - the script never reads or writes them.

    Returns a summary dict: {new, updated, removed, unchanged, total}.
    """
    if not api_records:
        raise ValueError("api_records is empty - nothing to merge.")

    api_fields = list(api_records[0].keys())
    if key_field not in api_fields:
        raise KeyError(f"Key field '{key_field}' not found in API data.")

    api_lookup = {str(r[key_field]): r for r in api_records}

    if os.path.exists(output_path):
        if backup:
            _backup_file(output_path)
        return _merge_into_existing(output_path, api_lookup, api_fields, key_field)
    else:
        return _create_new_workbook(output_path, api_lookup, api_fields, key_field)


# ---------------------------------------------------------------------------
# First run - create workbook from scratch
# ---------------------------------------------------------------------------

def _create_new_workbook(path, api_lookup, api_fields, key_field):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sessions"

    seed_names = [h["name"] for h in SEED_HUMAN_COLUMNS]
    all_cols = api_fields + [STATUS_COL, SYNCED_COL] + seed_names

    # ---- Header row ----
    for ci, name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if name in META_COLS:
            cell.fill = META_HEADER_FILL
            cell.font = META_HEADER_FONT
        elif name in seed_names:
            cell.fill = HUMAN_HEADER_FILL
            cell.font = HEADER_FONT
        else:
            cell.fill = API_HEADER_FILL
            cell.font = HEADER_FONT

    # ---- Data rows ----
    now_str = _now()
    status_ci = len(api_fields) + 1
    synced_ci = len(api_fields) + 2
    for ri, record in enumerate(api_lookup.values(), start=2):
        for ci, field in enumerate(api_fields, 1):
            _write_body_cell(ws, ri, ci, record.get(field, ""))
        _write_body_cell(ws, ri, status_ci, "New")
        _write_body_cell(ws, ri, synced_ci, now_str)

    # ---- Accept? dropdown ----
    accept_ci = next(
        (len(api_fields) + 3 + i for i, h in enumerate(SEED_HUMAN_COLUMNS) if h["name"] == "Accept?"),
        None,
    )
    if accept_ci:
        letter = get_column_letter(accept_ci)
        dv = DataValidation(type="list", formula1='"Yes,No,Maybe"', allow_blank=True, showDropDown=False)
        dv.error = "Please select Yes, No, or Maybe"
        dv.add(f"{letter}2:{letter}1048576")
        ws.add_data_validation(dv)

    # ---- Column widths ----
    for ci, field in enumerate(api_fields, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(field)) + 4, 14)
    ws.column_dimensions[get_column_letter(status_ci)].width = 14
    ws.column_dimensions[get_column_letter(synced_ci)].width = 18
    for i, h in enumerate(SEED_HUMAN_COLUMNS):
        ws.column_dimensions[get_column_letter(len(api_fields) + 3 + i)].width = h["width"]

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(api_lookup) + 1}"

    wb.save(path)
    return _stats(new=len(api_lookup))


# ---------------------------------------------------------------------------
# Subsequent runs - merge into existing workbook
# ---------------------------------------------------------------------------

def _merge_into_existing(path, api_lookup, api_fields, key_field):
    wb = load_workbook(path)
    ws = wb["Sessions"] if "Sessions" in wb.sheetnames else wb.active

    # Build header map: name -> 1-based column index
    col_map = {}
    for ci in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=ci).value
        if val:
            col_map[val] = ci

    key_ci = col_map.get(key_field)
    if key_ci is None:
        raise KeyError(f"Key column '{key_field}' not found in existing sheet.")

    status_ci = col_map.get(STATUS_COL)
    synced_ci = col_map.get(SYNCED_COL)

    # ONLY these columns will be written - everything else is hands-off.
    # We match by header name, so column position doesn't matter.
    api_col_map = {f: col_map[f] for f in api_fields if f in col_map}

    # Columns the script "owns" (for selective fill changes)
    owned_cis = set(api_col_map.values())
    if status_ci:
        owned_cis.add(status_ci)
    if synced_ci:
        owned_cis.add(synced_ci)

    # Index existing rows by key value
    existing = {}
    for ri in range(2, ws.max_row + 1):
        kv = str(ws.cell(row=ri, column=key_ci).value or "")
        if kv:
            existing[kv] = ri

    now_str = _now()
    stats = {"new": 0, "updated": 0, "removed": 0, "unchanged": 0}

    # ---- Update / flag existing rows ----
    for kv, ri in existing.items():
        if kv in api_lookup:
            record = api_lookup[kv]
            changed = False
            for field, ci in api_col_map.items():
                old = ws.cell(row=ri, column=ci).value
                new = record.get(field, "")
                if _norm(old) != _norm(new):
                    ws.cell(row=ri, column=ci, value=new)
                    changed = True

            if changed:
                _set_meta(ws, ri, status_ci, "Updated", synced_ci, now_str)
                # Clear removed highlighting on owned columns only
                for ci in owned_cis:
                    ws.cell(row=ri, column=ci).fill = NO_FILL
                stats["updated"] += 1
            else:
                prev = ws.cell(row=ri, column=status_ci).value if status_ci else None
                if prev == "Removed from API":
                    _set_meta(ws, ri, status_ci, "Restored", synced_ci, now_str)
                    for ci in owned_cis:
                        ws.cell(row=ri, column=ci).fill = NO_FILL
                else:
                    _set_meta(ws, ri, status_ci, "Unchanged", synced_ci, now_str)
                stats["unchanged"] += 1
        else:
            # Gone from API - highlight owned columns only, leave human cols alone
            _set_meta(ws, ri, status_ci, "Removed from API", synced_ci, now_str)
            for ci in owned_cis:
                ws.cell(row=ri, column=ci).fill = REMOVED_FILL
            stats["removed"] += 1

    # ---- Append new sessions (only into owned columns) ----
    next_ri = ws.max_row + 1
    for kv, record in api_lookup.items():
        if kv not in existing:
            for field, ci in api_col_map.items():
                cell = ws.cell(row=next_ri, column=ci, value=record.get(field, ""))
                cell.border = THIN_BORDER
                cell.font = BODY_FONT
                cell.fill = NEW_FILL
            _set_meta(ws, next_ri, status_ci, "New", synced_ci, now_str)
            stats["new"] += 1
            next_ri += 1

    stats["total"] = sum(stats.values())

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    wb.save(path)
    return stats


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(v):
    """Normalize a cell value for stable string comparison across runs."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v != v:          # NaN
            return ""
        if v == int(v):     # integer-valued float: 5.0 -> "5"
            return str(int(v))
    return str(v).replace("\r\n", "\n").replace("\r", "\n")

def _write_body_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.font = BODY_FONT

def _set_meta(ws, row, status_ci, status_val, synced_ci, synced_val):
    if status_ci:
        ws.cell(row=row, column=status_ci, value=status_val)
    if synced_ci:
        ws.cell(row=row, column=synced_ci, value=synced_val)

def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M")

def _stats(**kw):
    d = {"new": 0, "updated": 0, "removed": 0, "unchanged": 0}
    d.update(kw)
    d["total"] = sum(d.values())
    return d

def _backup_file(path):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    shutil.copy2(path, f"{base}_backup_{ts}{ext}")


# ---------------------------------------------------------------------------
# Demo - simulates first run, human edits, then a merge
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    OUT = "sessions_demo.xlsx"
    if os.path.exists(OUT):
        os.remove(OUT)

    run1 = [
        {"proposalguid": "AAA-111", "title": "Intro to Azure Functions", "speaker": "Alice Smith", "level": "100"},
        {"proposalguid": "BBB-222", "title": "Advanced Kubernetes", "speaker": "Bob Jones", "level": "300"},
        {"proposalguid": "CCC-333", "title": "DevOps Best Practices", "speaker": "Carol White", "level": "200"},
    ]

    print("=== Run 1 (fresh create) ===")
    print(merge_sessions(run1, OUT, backup=False))

    # Simulate a human opening the file and adding ratings + a custom column
    print("\n--- Simulating human edits ---")
    wb = load_workbook(OUT)
    ws = wb["Sessions"]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    ws.cell(2, hdr["Rating"], value=4)
    ws.cell(3, hdr["Rating"], value=5)
    ws.cell(2, hdr["Accept?"], value="Yes")
    ws.cell(3, hdr["Accept?"], value="Maybe")
    ws.cell(2, hdr["Reviewer Notes"], value="Great topic, love the demo angle")
    # Human adds a brand-new column the script has never seen
    new_col = ws.max_column + 1
    ws.cell(1, new_col, value="My Custom Flag")
    ws.cell(2, new_col, value="PRIORITY")
    ws.cell(3, new_col, value="BACKUP")
    wb.save(OUT)
    print("  Added ratings, notes, and a 'My Custom Flag' column.")

    # Run 2: API has changes - title updated, one removed, one added
    run2 = [
        {"proposalguid": "AAA-111", "title": "Intro to Azure Functions v2", "speaker": "Alice Smith", "level": "100"},
        {"proposalguid": "BBB-222", "title": "Advanced Kubernetes", "speaker": "Bob Jones", "level": "300"},
        # CCC-333 removed from API
        {"proposalguid": "DDD-444", "title": "GitHub Copilot Workshop", "speaker": "Dave Brown", "level": "200"},
    ]

    print("\n=== Run 2 (merge - human edits should survive) ===")
    print(merge_sessions(run2, OUT, backup=False))

    # Verify human data survived
    wb = load_workbook(OUT)
    ws = wb["Sessions"]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    print("\n--- Verification ---")
    for ri in range(2, ws.max_row + 1):
        guid = ws.cell(ri, hdr["proposalguid"]).value
        rating = ws.cell(ri, hdr["Rating"]).value
        accept = ws.cell(ri, hdr["Accept?"]).value
        notes = ws.cell(ri, hdr["Reviewer Notes"]).value
        flag = ws.cell(ri, hdr.get("My Custom Flag", 0)).value if "My Custom Flag" in hdr else None
        status = ws.cell(ri, hdr["_SyncStatus"]).value
        print(f"  {guid}: rating={rating}, accept={accept}, notes={notes!r}, flag={flag}, status={status}")

    print(f"\nOutput: {OUT}")
